import os
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from backend.config import OUTPUT_DIR

def save_to_excel(data, company_name):
    """
    Updates production data in the existing Excel file.
    
    1. Searches column B (starting at row 5) for company_name.
    2. Uses the fixed header in row 4—quarters in columns C to AT and years in columns AV to BF—to map data.
    3. If the company exists, updates its production figures in the corresponding columns.
    4. If the company does not exist, appends a new row at the bottom, placing company_name in column B and filling in production figures.
    5. Preserves existing formatting.
    """
    excel_path = os.path.join(OUTPUT_DIR, "lithium_production.xlsx")
    wb = load_workbook(excel_path)
    ws = wb["Lithium Production"]

    # Define fixed column ranges based on setup_excel:
    # Quarter labels are in row 4, columns C to AT.
    quarter_start = "C"
    quarter_end = "AT"
    # Year labels are in row 4, columns AV to BF.
    year_start = "AV"
    year_end = "BF"

    q_start_idx = column_index_from_string(quarter_start)
    q_end_idx = column_index_from_string(quarter_end)
    y_start_idx = column_index_from_string(year_start)
    y_end_idx = column_index_from_string(year_end)

    # Build a mapping from header label to column index (for both quarters and years)
    label_to_col = {}

    # For quarters:
    for col in range(q_start_idx, q_end_idx + 1):
        header = ws.cell(row=4, column=col).value
        if header is not None:
            label_to_col[str(header).strip()] = col

    # For years:
    for col in range(y_start_idx, y_end_idx + 1):
        header = ws.cell(row=4, column=col).value
        if header is not None:
            label_to_col[str(header).strip()] = col

    # Look for the company name in column B, starting at row 5
    company_row = None
    for row in range(5, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=2).value  # Column B
        if cell_value and str(cell_value).strip().lower() == company_name.strip().lower():
            company_row = row
            break

    if company_row is None:
        # If not found, append a new row at the bottom
        company_row = ws.max_row + 1
        ws.cell(row=company_row, column=2).value = company_name

    # Update production figures in the matching columns
    for key, value in data.items():
        label = key.strip()
        if label in label_to_col:
            col_idx = label_to_col[label]
            ws.cell(row=company_row, column=col_idx).value = value
        else:
            print(f"Warning: Label '{label}' not found in header row.")

    # Save workbook (this preserves the formatting already in the file)
    wb.save(excel_path)
    print(f"✅ Data updated for '{company_name}' in {excel_path}")
    return excel_path
